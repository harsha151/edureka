import pandas as pd
from openpyxl import Workbook, load_workbook

# Function to process user input (replace with your actual logic)
def process_input(user_input):
  # Access data from Excel file (assuming "chat_bot.xlsx" exists)
  try:
    data = pd.read_excel("give.xlsx")
    responses = data.to_dict(orient="records")  # Convert to list of dictionaries
  except FileNotFoundError:
    print("Error: Excel file 'give.xlsx' not found.")
    return None  # Indicate error or provide a default response

  # Process user input based on data
  for entry in responses:
    if user_input in entry["Input"]:
      return entry["Response"]
  return "Sorry, I didn't understand. Can you rephrase?"

# Run the bot
while True:
  user_input = input("You: ")
  response = process_input(user_input)

  # Log conversation (optional)
  if response:
    # Open existing workbook in append mode (if it exists)
    try:
      wb = load_workbook("update.xlsx", read_only=False)  # Open for modification
    except FileNotFoundError:
      # Create a new workbook if file doesn't exist
      wb = Workbook()
      wb.active.title = "Conversation Log"  # Set sheet title

    sheet = wb.active
    sheet.append([user_input, response])  # Add conversation data to a new row

    # Save the changes to the existing file
    wb.save("update.xlsx")

  if response:
    print("Bot: " + response)
  else:
    print(response)  # Print error message or default response

  # Exit loop with 'quit' command (optional)
  if user_input.lower() == "quit":
    break
